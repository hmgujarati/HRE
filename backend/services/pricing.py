"""Pricing services — price-history audit, bulk-discount apply, Excel parsing.

These are stateless internal helpers used by `routers/variants.py`,
`routers/families.py` (variant Excel upload), and `routers/pricing.py`
(bulk-discount + price Excel import). Lives outside `routers/` so it can be
shared without creating per-router circular dependencies.
"""

from __future__ import annotations
import uuid
from io import BytesIO
from typing import Any, Dict, List, Tuple

import openpyxl
from fastapi import HTTPException

from core import calc_final_price, db, now_iso


# ─────────────────── Price History ───────────────────

async def record_price_history(variant_before: Dict[str, Any] | None,
                                variant_after: Dict[str, Any],
                                changed_by: str, reason: str) -> None:
    """Audit-trail every price-affecting change. `variant_before=None` means a
    creation event with no prior state, so we skip. `{}` means a creation
    event WITH a baseline row to record."""
    if variant_before is None:
        return
    fields = ["base_price", "discount_percentage", "manual_price",
              "manual_price_override", "final_price"]
    if variant_before and not any(variant_before.get(f) != variant_after.get(f) for f in fields):
        return
    entry = {
        "id": str(uuid.uuid4()),
        "product_variant_id": variant_after["id"],
        "product_family_id": variant_after.get("product_family_id"),
        "changed_by": changed_by,
        "old_base_price": variant_before.get("base_price"),
        "new_base_price": variant_after.get("base_price"),
        "old_discount_percentage": variant_before.get("discount_percentage"),
        "new_discount_percentage": variant_after.get("discount_percentage"),
        "old_manual_price": variant_before.get("manual_price"),
        "new_manual_price": variant_after.get("manual_price"),
        "old_manual_price_override": variant_before.get("manual_price_override"),
        "new_manual_price_override": variant_after.get("manual_price_override"),
        "old_final_price": variant_before.get("final_price"),
        "new_final_price": variant_after.get("final_price"),
        "change_reason": reason,
        "changed_at": now_iso(),
    }
    await db.price_history.insert_one(entry.copy())


# ─────────────────── Bulk Discount ───────────────────

async def apply_bulk_discount(query: Dict[str, Any], discount: float,
                               user_email: str, reason: str) -> int:
    variants = await db.product_variants.find(query, {"_id": 0}).to_list(10000)
    count = 0
    for v in variants:
        before = dict(v)
        new_doc = dict(v)
        new_doc["discount_percentage"] = float(discount)
        new_doc["final_price"] = calc_final_price(
            new_doc["base_price"], new_doc["discount_percentage"],
            new_doc.get("manual_price_override", False), new_doc.get("manual_price"),
        )
        new_doc["updated_at"] = now_iso()
        await db.product_variants.update_one({"id": v["id"]}, {"$set": {
            "discount_percentage": new_doc["discount_percentage"],
            "final_price": new_doc["final_price"],
            "updated_at": new_doc["updated_at"],
        }})
        await record_price_history(before, new_doc, user_email, reason)
        count += 1
    return count


# ─────────────────── Excel parsing helpers ───────────────────

def norm(s: Any) -> str:
    return str(s or "").strip().lower().replace(".", "").replace(" ", "")


def norm_code(s: Any) -> str:
    """Normalise product code for matching: uppercase, remove all whitespace."""
    return "".join(str(s or "").upper().split())


def is_number(v: Any) -> bool:
    if v is None:
        return False
    try:
        float(str(v).strip().replace(",", ""))
        return True
    except Exception:
        return False


def parse_variant_workbook(content: bytes) -> Tuple[List[str], List[List[Any]]]:
    """Return (header_keys, data_rows). Tolerates 1/2/3-row merged headers
    (e.g. row 0 has 'CABLE SIZE / HOLE E / DIMENSIONS / PROD. CODE' with merged
    cells, row 1 is empty due to vertical merges, row 2 carries dimension
    sub-keys A/B/C/D...)."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        raise HTTPException(status_code=400, detail="Empty workbook")
    n_cols = max(len(r) for r in rows)
    rows = [list(r) + [None] * (n_cols - len(r)) for r in rows]

    def looks_like_code(v):
        if v is None:
            return False
        s = str(v).strip()
        return bool(s) and any(ch.isdigit() for ch in s) and any(ch.isalpha() for ch in s)

    def looks_like_number(v):
        if v is None:
            return False
        try:
            float(str(v).strip())
            return True
        except Exception:
            return False

    data_start = None
    for i, row in enumerate(rows):
        if any(looks_like_code(c) for c in row) and sum(1 for c in row if looks_like_number(c)) >= 3:
            data_start = i
            break
    if data_start is None or data_start == 0:
        raise HTTPException(status_code=400, detail="Could not locate data rows in spreadsheet")

    GENERIC = {"dimensions", "dimension", "specs", "specification"}
    headers: List[str] = []
    for ci in range(n_cols):
        chosen = ""
        for ri in range(data_start):
            v = rows[ri][ci]
            if v is None:
                continue
            sv = str(v).strip()
            if not sv:
                continue
            if sv.lower() in GENERIC:
                continue
            chosen = sv
        headers.append(chosen)
    data = rows[data_start:]
    data = [r for r in data if any(c is not None and str(c).strip() != "" for c in r)]
    return headers, data


def classify_header(h: str) -> str:
    """Return 'cable'|'hole'|'code'|'price'|'dim:<key>'|'skip'."""
    n = norm(h)
    if not n:
        return "skip"
    if "prod" in n or "code" in n:
        return "code"
    if "cable" in n or n == "mm2" or ("size" in n and "hole" not in n):
        return "cable"
    if "hole" in n or n == "e":
        return "hole"
    if any(k in n for k in ["price", "rate", "mrp", "cost", "amount", "hre"]):
        return "price"
    return f"dim:{h.strip()}"


def parse_price_workbook(content: bytes) -> Tuple[List[str], List[List[Any]]]:
    """Lightweight parser for a 2-column-style price sheet:
    a Product Code column + a Price/Rate/HRE column. Used as a fallback when
    `parse_variant_workbook` can't find a 3+ numeric data row (the variant
    sheet shape). Tolerates a header row anywhere in the first 5 rows and
    extra empty trailing columns."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        raise HTTPException(status_code=400, detail="Empty workbook")
    n_cols = max((len(r) for r in rows), default=0)
    rows = [list(r) + [None] * (n_cols - len(r)) for r in rows]

    # Find header row: at least one cell maps to 'code' and at least one to 'price'.
    header_row_idx = None
    for ri in range(min(5, len(rows))):
        roles = [classify_header(str(c)) if c is not None else "skip" for c in rows[ri]]
        if "code" in roles and "price" in roles:
            header_row_idx = ri
            break
    if header_row_idx is None:
        raise HTTPException(
            status_code=400,
            detail="Could not find a header row with both a Product Code and a Price column. "
                   "Ensure the first row has labels like 'PROD. CODE' and 'Price'/'Rate'/'HRE'.",
        )
    headers = [str(c).strip() if c is not None else "" for c in rows[header_row_idx]]
    data = rows[header_row_idx + 1:]
    data = [r for r in data if any(c is not None and str(c).strip() != "" for c in r)]
    return headers, data
